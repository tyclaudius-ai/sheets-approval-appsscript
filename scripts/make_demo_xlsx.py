#!/usr/bin/env python3
"""Generate a client-ready demo XLSX for the Sheets approvals microproduct.

Why XLSX?
- Lets someone download a single file, upload to Google Drive, and open as a Google Sheet.
- Includes the correct tab names + headers + a couple sample rows.

This does NOT include the Apps Script (that still needs to be added via Extensions → Apps Script).
"""

from __future__ import annotations

from pathlib import Path
import csv
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

ROOT = Path(__file__).resolve().parents[1]
DEMO_DIR = ROOT / "demo"
OUT_PATH = DEMO_DIR / "Sheets-Approvals-Demo.xlsx"


def read_csv(path: Path) -> list[list[str]]:
    rows: list[list[str]] = []
    with path.open("r", newline="", encoding="utf-8") as f:
        for row in csv.reader(f):
            rows.append(row)
    return rows


def style_header_row(ws, header_row: int = 1):
    header_fill = PatternFill("solid", fgColor="1F2937")  # slate-800
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[header_row]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="center")


def autosize_columns(ws, max_width: int = 60, min_width: int = 10):
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for c in col:
            v = "" if c.value is None else str(c.value)
            max_len = max(max_len, len(v))
        ws.column_dimensions[col_letter].width = max(min_width, min(max_width, max_len + 2))


def main() -> None:
    wb = Workbook()

    # Make output reproducible: Excel metadata timestamps can otherwise change on each run.
    fixed_dt = datetime(2000, 1, 1, 0, 0, 0)
    wb.properties.created = fixed_dt
    wb.properties.modified = fixed_dt
    wb.properties.creator = "sheets-approval-appsscript"
    wb.properties.lastModifiedBy = "sheets-approval-appsscript"

    # Requests
    ws_req = wb.active
    ws_req.title = "Requests"
    req_rows = read_csv(DEMO_DIR / "Requests.csv")
    for r in req_rows:
        ws_req.append(r)
    style_header_row(ws_req, 1)
    ws_req.freeze_panes = "A2"
    autosize_columns(ws_req)

    # Audit
    ws_audit = wb.create_sheet("Audit")
    audit_rows = read_csv(DEMO_DIR / "Audit.csv")
    for r in audit_rows:
        ws_audit.append(r)
    style_header_row(ws_audit, 1)
    ws_audit.freeze_panes = "A2"
    autosize_columns(ws_audit)

    # README tab for humans
    ws_readme = wb.create_sheet("README")
    ws_readme.append(["Sheets Approvals + Audit Trail — Demo Workbook"])
    ws_readme.append([""])
    ws_readme.append([
        "1) Upload this .xlsx to Google Drive and open it as a Google Sheet."])
    ws_readme.append([
        "2) Extensions → Apps Script: paste Code.gs from the repo release bundle, then save."])
    ws_readme.append([
        "3) Reload the sheet → Approvals menu appears → run 'Create demo setup' or approve/reject rows."])
    ws_readme.append([""])
    ws_readme.append([
        "Repo: https://github.com/tyclaudius-ai/sheets-approval-appsscript"])
    ws_readme["A1"].font = Font(bold=True, size=14)
    ws_readme.column_dimensions["A"].width = 100

    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_PATH)
    print(f"Wrote: {OUT_PATH}")


if __name__ == "__main__":
    main()
